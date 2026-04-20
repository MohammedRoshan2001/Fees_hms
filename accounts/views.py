# =======================
# Django Core Imports
# =======================
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.urls import reverse
from django.views import generic, View
from django.db import transaction
from django.db.models import Sum, Prefetch
from django.core.paginator import Paginator

# =======================
# Auth Imports
# =======================
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User

# =======================
# Third-party Imports
# =======================
import openpyxl
from openpyxl.styles import Font
from datetime import datetime

# =======================
# Local App Models
# =======================
from .models import *
from accounts.models import *

# =======================
# Local App Forms
# =======================
from .forms import *

class Abst:
    def __init__(self,sl_no,dept,year,demand,collection,balance,count):
        self.sl_no=sl_no 
        self.dept=dept 
        self.year=year
        self.demand=demand 
        self.collection=collection 
        self.balance=balance 
        self.count=count
#Creating our view, it is a class based view 
class Collection:
    def __init__(self,receipt_no,date,fees) -> None:
        self.receipt_no=receipt_no 
        self.date=date 
        self.fees=fees
        pass
class Fees_Collection:
    def __init__(self,student,history) -> None:
        self.student=student 
        self.history=history
        pass

class Sub_Category:
    def __init__(self,name,count) -> None:
        self.name = name
        self.count = count
        pass

 
def reciept(request):
    return render(request,'print_reciept.html')
    pass


def index(request):
 
    return render(request, 'base.html')


def add_student(request):
    if request.method=="POST":
        usn=request.POST.get('usn')
        year=request.POST.get('year')
        repeater = request.POST.get('repeater')
        print(repeater, type(repeater))
        print(usn)
        print(year)
        st=Student.objects.get(roll_no2=usn)
        print(st)
        y=int(year)
        rep = False
        if repeater == "1":
            rep = True

        student_objects=Fees_Details.objects.filter(student=st,year=y)
        if rep == True and len(student_objects) ==0:
            return HttpResponse('<h1> Given Student is not a Repeater </h1>') 
        if rep == False and len(student_objects) ==1:
            return render(request,'student_exist.html')
        if len(student_objects) == 2:
            return HttpResponse('<h1> Maximum attempts completed for the given year </h1>')
        academic_year1=request.POST.get('academic_year')
        academic_year=Academic_Year.objects.get(academic_year=academic_year1)
        student_academic_year=Fees_Details.objects.filter(student=st,academic_year=academic_year)
        if len(student_academic_year)>0:
            return HttpResponse('<h1> Student and acadmeic year combination already exists  </h1>')
   
        years_completed = 0
        for i in Fees_Details.objects.filter(student = st):
            years_completed = max(years_completed,i.year)
            
        if rep == True:
            years_completed -=1

        try:
            fees_structure = Fees_Structure.objects.get(
                year=y,
                academic_year=academic_year,
                category=st.category,
                repeater=rep,
                is_lateral=st.Is_lateral
            )
        except Fees_Structure.DoesNotExist:
            return HttpResponse(
                f"<h1>Fees Structure does not exist for "
                f"Year: {y}, Academic Year: {academic_year}, "
                f"Category: {st.category}, Repeater: {rep}, "
                f"Is Lateral: {st.Is_lateral}</h1>"
            )
        fees_obj=Fees_Details(student=st,year=y,repeater = rep,academic_year=academic_year,total_fees=fees_structure.total_fees,balance=fees_structure.total_fees)
        st.year_completed = years_completed
        fees_obj.save()
        st.save()
        return HttpResponseRedirect(reverse('success'))
    else:
        student_list=Student.objects.all()
        year_list=Academic_Year.objects.all().order_by('-academic_year')
        context={'student_list':student_list,'year_list':year_list}
        return render(request,'add_student.html',context)
    pass



def fees_updation(request):

    # ✅ GET instead of POST
    dept = request.GET.get('dept') or "---"
    year = request.GET.get('year') or "---"
    academic_year = request.GET.get('academic_year') or "---"
    reg_no = (request.GET.get('reg_no') or "").strip().upper()
    bal = request.GET.get('bal')
    is_lateral = request.GET.get('is_lateral') == '1'
    display_alphabetically = request.GET.get('display_alphabetically') == '1'

    # Base queryset
    fees_qs = Fees_Details.objects.filter(
        student__cancel_admission=False
    ).select_related('student', 'academic_year')

    # ✅ Filters
    if dept != "---":
        fees_qs = fees_qs.filter(student__dep=dept)

    if year != "---":
        try:
            fees_qs = fees_qs.filter(year=int(year))
        except:
            pass

    if academic_year != "---":
        fees_qs = fees_qs.filter(academic_year__academic_year=academic_year)

    if reg_no:
        fees_qs = fees_qs.filter(student__roll_no2__iexact=reg_no)

    if bal:
        fees_qs = fees_qs.filter(balance__gt=0)

    if is_lateral:
        fees_qs = fees_qs.filter(student__Is_lateral=True)

    # ✅ Sorting
    if display_alphabetically:
        fees_qs = fees_qs.order_by('student__name')
    else:
        fees_qs = fees_qs.order_by('-id')

    # ✅ Pagination
    paginator = Paginator(fees_qs, 30)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # ✅ Totals
    totals = fees_qs.aggregate(
        total=Sum('total_fees'),
        collection=Sum('collection'),
        balance=Sum('balance')
    )

    # ✅ Build page list
    student_list = []
    for i in page_obj:
        history = [
            Collection(j.fees_receipt_no, j.date, j.total_fees)
            for j in History.objects.filter(
                student=i.student,
                year=i.year,
                academic_year=i.academic_year
            )
        ]
        student_list.append(Fees_Collection(i, history))

    context = {
        'student_list': student_list,
        'page_obj': page_obj,
        'year_list': Academic_Year.objects.all().order_by('-academic_year'),
        'dep_list': department_choices,
        'total': totals['total'] or 0,
        'collection': totals['collection'] or 0,
        'balance': totals['balance'] or 0,
        'dept': dept,
        'only_balance': bal,
        'academic_year': academic_year,
        'reg_no': reg_no,
        'year': year,
        'is_lateral': is_lateral,
        'display_alphabetically': display_alphabetically,
    }

    return render(request, 'fees_updation.html', context)

@login_required
def export_fees_excel(request):

    dept = request.GET.get('dept') or "---"
    year = request.GET.get('year') or "---"
    academic_year = request.GET.get('academic_year') or "---"
    reg_no = (request.GET.get('reg_no') or "").strip().upper()
    bal = request.GET.get('bal')
    is_lateral = request.GET.get('is_lateral')

    fees_qs = Fees_Details.objects.filter(
        student__cancel_admission=False
    ).select_related('student', 'academic_year')

    # =========================
    # FILTERS
    # =========================
    if dept != "---":
        fees_qs = fees_qs.filter(student__dep=dept)

    if year != "---":
        try:
            fees_qs = fees_qs.filter(year=int(year))
        except:
            pass

    if academic_year != "---":
        fees_qs = fees_qs.filter(academic_year__academic_year=academic_year)

    if reg_no:
        fees_qs = fees_qs.filter(student__roll_no2__iexact=reg_no)

    if bal:
        fees_qs = fees_qs.filter(balance__gt=0)

    if is_lateral == '1':
        fees_qs = fees_qs.filter(student__Is_lateral=True)

    # =========================
    # GRAND TOTAL (FAST DB WAY)
    # =========================
    grand_totals = fees_qs.aggregate(
        total_fees_sum=Sum('total_fees'),
        collection_sum=Sum('collection'),
        balance_sum=Sum('balance')
    )

    # =========================
    # WORKBOOK
    # =========================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fees Report"

    headers = [
        "Sl.No", "Name", "Reg No", "Dept", "Category",
        "Year", "Academic Year", "Total Fees",
        "Collection", "Balance", "Repeatation", "Status", "History"
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    sl_no = 1

    for i in fees_qs:

        ws.append([
            sl_no,
            i.student.name,
            i.student.roll_no2,
            i.student.dep,
            i.student.category,
            i.year,
            i.academic_year.academic_year,
            i.total_fees,
            i.collection,
            i.balance,
            "Yes" if i.repeater else "",
            "Detained" if i.is_detained else "Completed",
            ""
        ])

        history = History.objects.filter(
            student=i.student,
            year=i.year,
            academic_year=i.academic_year
        )

        if history.exists():

            ws.append([
                "", "", "", "", "", "", "",
                "Receipt No", "Date", "Fees",
                "", "", ""
            ])

            for j in history:
                ws.append([
                    "", "", "", "", "", "", "",
                    j.fees_receipt_no,
                    str(j.date),
                    j.total_fees,
                    "", "", ""
                ])

        sl_no += 1

    # =========================
    # GRAND TOTAL ROW
    # =========================
    ws.append([])

    ws.append([
        "GRAND TOTAL", "", "", "", "", "", "",
        grand_totals['total_fees_sum'] or 0,
        grand_totals['collection_sum'] or 0,
        grand_totals['balance_sum'] or 0,
        "", "", ""
    ])

    # =========================
    # AUTO WIDTH
    # =========================
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    # =========================
    # RESPONSE
    # =========================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    timestamp = datetime.now().strftime("%d-%m-%Y---%H-%M-%S")
    filename = f"fees_report__{timestamp}.xlsx"

    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response

@login_required
def fees_updation_summary(request):

    # ✅ GET instead of POST (pagination friendly)
    dept = request.GET.get('dept') or "---"
    year = request.GET.get('year') or "---"
    academic_year = request.GET.get('academic_year') or "---"
    reg_no = (request.GET.get('reg_no') or "").strip().upper()
    bal = request.GET.get('bal')
    is_lateral = request.GET.get('is_lateral') == '1'
    is_snq = request.GET.get('is_snq') == '1'
    display_alphabetically = request.GET.get('display_alphabetically') == '1'

    year_filter = -1 if year == "---" else int(year)

    # ✅ BASE QUERY (optimized)
    fees_qs = Fees_Details.objects.select_related(
        'student', 'academic_year'
    ).filter(
        student__cancel_admission=False
    ).exclude(year=0)

    # ✅ FILTERS
    if dept != "---":
        fees_qs = fees_qs.filter(student__dep=dept)

    if year_filter != -1:
        fees_qs = fees_qs.filter(year=year_filter)

    if academic_year != "---":
        fees_qs = fees_qs.filter(academic_year__academic_year=academic_year)

    if reg_no:
        fees_qs = fees_qs.filter(student__roll_no2__iexact=reg_no)

    if bal:
        fees_qs = fees_qs.filter(balance__gt=0)

    if is_lateral:
        fees_qs = fees_qs.filter(student__Is_lateral=True)

    if is_snq:
        fees_qs = fees_qs.filter(student__category='SNQ')

    # ✅ ORDERING
    if display_alphabetically:
        fees_qs = fees_qs.order_by('student__name')
    else:
        fees_qs = fees_qs.order_by('-id')

    # ✅ TOTALS (DB LEVEL 🔥)
    totals = fees_qs.aggregate(
        total=Sum('total_fees'),
        collection=Sum('collection'),
        balance=Sum('balance')
    )

    # ✅ PAGINATION (DB slicing 🔥)
    paginator = Paginator(fees_qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # ✅ PREFETCH HISTORY (no N+1)
    history_qs = History.objects.all()
    page_obj.object_list = page_obj.object_list.prefetch_related(
        Prefetch('student__history_set', queryset=history_qs)
    )

    student_list = []
    for i in page_obj:
        history = [
            Collection(j.fees_receipt_no, j.date, j.total_fees)
            for j in i.student.history_set.all()
            if j.year == i.year
        ]
        student_list.append(Fees_Collection(i, history))

    context = {
        'student_list': student_list,
        'page_obj': page_obj,
        'year_list': Academic_Year.objects.all().order_by('-academic_year'),
        'dep_list': department_choices,

        # totals
        'total': totals['total'] or 0,
        'collection': totals['collection'] or 0,
        'balance': totals['balance'] or 0,

        # filters
        'dept': dept,
        'year': year,
        'academic_year': academic_year,
        'reg_no': reg_no,
        'only_balance': bal,
        'is_lateral': is_lateral,
        'is_snq': is_snq,
        'display_alphabetically': display_alphabetically,
    }

    return render(request, 'fees_updation_summary.html', context)
    
    pass

@login_required
def export_fees_excel(request):

    dept = request.GET.get('dept') or "---"
    year = request.GET.get('year') or "---"
    academic_year = request.GET.get('academic_year') or "---"
    reg_no = (request.GET.get('reg_no') or "").strip().upper()
    bal = request.GET.get('bal')
    is_lateral = request.GET.get('is_lateral') == '1'
    is_snq = request.GET.get('is_snq') == '1'

    year_filter = -1 if year == "---" else int(year)

    qs = Fees_Details.objects.select_related('student', 'academic_year')\
        .filter(student__cancel_admission=False)\
        .exclude(year=0)

    # ✅ APPLY FILTERS
    if dept != "---":
        qs = qs.filter(student__dep=dept)

    if year_filter != -1:
        qs = qs.filter(year=year_filter)

    if academic_year != "---":
        qs = qs.filter(academic_year__academic_year=academic_year)

    if reg_no:
        qs = qs.filter(student__roll_no2__iexact=reg_no)

    if bal:
        qs = qs.filter(balance__gt=0)

    if is_lateral:
        qs = qs.filter(student__Is_lateral=True)

    if is_snq:
        qs = qs.filter(student__category='SNQ')

    # ✅ DB LEVEL TOTALS (FAST 🔥)
    totals = qs.aggregate(
        total=Sum('total_fees'),
        collection=Sum('collection'),
        balance=Sum('balance')
    )

    total_val = totals['total'] or 0
    collection_val = totals['collection'] or 0
    balance_val = totals['balance'] or 0

    # ✅ EXCEL CREATION
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fees Summary"

    bold_font = Font(bold=True)

    # HEADER
    headers = [
        "Sl.No", "Name", "Reg No", "Dept", "Category",
        "Year", "Academic Year", "Total Fees",
        "Collection", "Balance"
    ]

    ws.append(headers)

    # make header bold
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = bold_font

    # DATA
    for idx, i in enumerate(qs.iterator(), start=1):  # memory safe
        ws.append([
            idx,
            i.student.name,
            i.student.roll_no2,
            i.student.dep,
            i.student.category,
            i.year,
            i.academic_year.academic_year,
            i.total_fees,
            i.collection,
            i.balance
        ])

    # ✅ EMPTY ROW
    ws.append([])

    # ✅ TOTAL HEADER ROW
    total_row_index = ws.max_row + 1
    ws.append([
        "", "", "", "", "", "",
        "Grand Total",
        "Collection",
        "Balance",
        ""
    ])

    for col in range(1, 11):
        ws.cell(row=total_row_index, column=col).font = bold_font

    # ✅ TOTAL VALUES ROW
    value_row_index = ws.max_row + 1
    ws.append([
        "", "", "", "", "", "",
        total_val,
        collection_val,
        balance_val,
        ""
    ])

    for col in range(1, 11):
        ws.cell(row=value_row_index, column=col).font = bold_font

    # ✅ RESPONSE
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=fees_summary.xlsx'

    wb.save(response)
    return response

@login_required
def update_form(request,roll_no,year,academic_year):
    if not request.user.is_superuser:
        return HttpResponse("<h1>Permission denied!!</h1>")
    student=Student.objects.get(roll_no2=roll_no)
    fees=Fees_Details.objects.get(student=student,year=year,academic_year=academic_year)
    if request.method=='POST':
        roll_no=request.POST.get('roll_no')
        year=request.POST.get('year')
        amount=request.POST.get('amount')
        receipt_no=request.POST.get('receipt_no')
        receipt_no=int(receipt_no)
        date=Date.objects.get(pk="1").date
        try:
            fees_structure = Fees_Structure.objects.get(
                year=int(year),
                academic_year=academic_year,
                category=student.category,
                repeater=fees.repeater,
                is_lateral=student.Is_lateral
            )
        except Fees_Structure.DoesNotExist:
            return HttpResponse(
                f"<h1>Fees Structure does not exist for "
                f"Year: {year}, Academic Year: {academic_year}, "
                f"Category: {student.category}, Repeater: {fees.repeater}, "
                f"Is Lateral: {student.Is_lateral}</h1>"
            )
        try:
            hist = History.objects.get(pk=receipt_no)
            return render(request,'fee_reciept_exist.html')
        except:
            year=int(year)
            amount=int(amount)
            threshold = sum([
                fees_structure.admission_fees,
                fees_structure.id_fees,
                fees_structure.management_fees,
                fees_structure.lib_fees,
                fees_structure.assn_fees,
                fees_structure.rr_fees,
                fees_structure.swf_fees,
                fees_structure.twf_fees,
                fees_structure.lab_fees,
                fees_structure.sp_fees,
                fees_structure.nss_fees,
                fees_structure.dev_fees
            ])
            context={}
                
            if fees.collection<threshold:
                if amount>=threshold:
                    tution_fees=amount-threshold
                    hist=History(pk=receipt_no,student=student,year=fees.year,academic_year=fees.academic_year.academic_year,tution_fees=tution_fees,
                            admission_fees=fees_structure.admission_fees,id_fees=fees_structure.id_fees,management_fees=fees_structure.management_fees,lib_fees=fees_structure.lib_fees,assn_fees=fees_structure.assn_fees,
                            rr_fees=fees_structure.rr_fees,swf_fees=fees_structure.swf_fees,twf_fees=fees_structure.twf_fees,
                            lab_fees=fees_structure.lab_fees,sp_fees=fees_structure.sp_fees,nss_fees=fees_structure.nss_fees,dev_fees=fees_structure.dev_fees,date=date)
                    hist.total_fees=hist.tution_fees+ hist.admission_fees + hist.id_fees + hist.management_fees + hist.lib_fees + hist.assn_fees + hist.rr_fees + hist.swf_fees + hist.twf_fees + hist.lab_fees+ hist.sp_fees+ hist.nss_fees+ hist.dev_fees
                    hist.save()
                    fees.collection+=hist.total_fees
                    fees.balance=fees.total_fees-fees.collection
                    fees.save()
                    if fees.balance==0:
                        fees.student.year_completed+=1
                        fees.student.save()
                    context['hist']=hist
                    return render(request,'print_reciept.html',context)
                else:
                    hist=History(fees_receipt_no=receipt_no,student=student,year=fees.year,academic_year=fees.academic_year.academic_year,tution_fees=amount,date=date)
                    hist.total_fees=hist.tution_fees+ hist.admission_fees + hist.id_fees + hist.management_fees + hist.lib_fees + hist.assn_fees + hist.rr_fees + hist.swf_fees + hist.twf_fees + hist.lab_fees+ hist.sp_fees+ hist.nss_fees+ hist.dev_fees
                    hist.save()
                    fees.collection+=hist.total_fees
                    fees.balance=fees.total_fees-fees.collection
                    fees.save()
                    if fees.balance==0:
                        fees.student.year_completed+=1
                        fees.student.save()
                    context['hist']=hist
                    return render(request,'print_reciept.html',context)
            else:
                hist=History(fees_receipt_no=receipt_no,student=student,year=fees.year,academic_year=fees.academic_year.academic_year,tution_fees=amount,date=date)
                hist.total_fees=hist.tution_fees+ hist.admission_fees + hist.id_fees + hist.management_fees + hist.lib_fees + hist.assn_fees + hist.rr_fees + hist.swf_fees + hist.twf_fees + hist.lab_fees+ hist.sp_fees+ hist.nss_fees+ hist.dev_fees
                hist.save()
                fees.collection+=hist.total_fees
                fees.balance=fees.total_fees-fees.collection
                fees.save()
                if fees.balance==0:
                    fees.student.year_completed+=1
                    fees.student.save()
                context['hist']=hist
            
                return render(request,'print_reciept.html',context)         
    else:            
        context={
            'roll_no':roll_no,
            'year':year,
            'fees':fees,
            'student':student,
            'date':Date.objects.get(pk="1")
        }
        return render(request,'update_form.html',context)
       
    
       
    

def success(request):
    return render(request, 'success.html')



FIELDS = [
    'tution_fees','admission_fees','id_fees','management_fees',
    'lib_fees','assn_fees','rr_fees','swf_fees','twf_fees',
    'lab_fees','sp_fees','nss_fees','dev_fees','total_fees'
]


# =========================
# ✅ DAY HISTORY VIEW
# =========================
@login_required
def day_history(request):

    from_date = request.POST.get('from_date')
    to_date = request.POST.get('to_date')

    qs = History.objects.all().order_by('date', 'fees_receipt_no')

    if from_date and to_date:
        qs = qs.filter(date__range=[from_date, to_date])

    # ✅ -------- GRAND TOTAL (FULL DATA, NOT PAGINATED) --------
    grand_totals = {k: 0 for k in FIELDS}

    for obj in qs:
        for key in FIELDS:
            grand_totals[key] += getattr(obj, key) or 0

    # ✅ -------- PAGINATION --------
    paginator = Paginator(qs, 300)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    page_list = list(page_obj)

    # ✅ -------- GROUPING --------
    final_list = []

    if page_list:
        min_fees_no = page_list[0].fees_receipt_no
        max_fees_no = page_list[0].fees_receipt_no
        current_date = page_list[0].date

        totals = {k: 0 for k in FIELDS}

        for obj in page_list:

            if obj.date != current_date:
                final_list.append(History(
                    fees_receipt_no=f"{min_fees_no}-{max_fees_no}",
                    date=current_date,
                    **totals
                ))

                # reset
                min_fees_no = obj.fees_receipt_no
                max_fees_no = obj.fees_receipt_no
                totals = {k: 0 for k in FIELDS}
                current_date = obj.date

            else:
                max_fees_no = obj.fees_receipt_no

            for key in FIELDS:
                totals[key] += getattr(obj, key) or 0

        # last group
        final_list.append(History(
            fees_receipt_no=f"{min_fees_no}-{max_fees_no}",
            date=current_date,
            **totals
        ))

    context = {
        'day_list': final_list,
        'page_obj': page_obj,
        'from_date': from_date or '',
        'to_date': to_date or '',
        'grand_totals': grand_totals
    }

    return render(request, 'day_history.html', context)


# =========================
# ✅ EXCEL EXPORT
# =========================
@login_required
def day_history_excel(request):

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    qs = History.objects.all().order_by('date', 'fees_receipt_no')

    if from_date and to_date:
        qs = qs.filter(date__range=[from_date, to_date])

    qs = list(qs)

    wb = openpyxl.Workbook()
    ws = wb.active

    headers = [
        'Date','Receipt Range','Total','Tution','Admission','ID',
        'Management','Library','Assn','RR','SWF','TWF','Lab','SP','NSS','Dev'
    ]
    ws.append(headers)

    # ✅ GRAND TOTAL INIT
    grand_totals = {k: 0 for k in FIELDS}

    if qs:
        min_fees_no = qs[0].fees_receipt_no
        max_fees_no = qs[0].fees_receipt_no
        current_date = qs[0].date

        totals = {k: 0 for k in FIELDS}

        for obj in qs:

            if obj.date != current_date:
                ws.append([
                    current_date,
                    f"{min_fees_no}-{max_fees_no}",
                    totals['total_fees'],
                    totals['tution_fees'],
                    totals['admission_fees'],
                    totals['id_fees'],
                    totals['management_fees'],
                    totals['lib_fees'],
                    totals['assn_fees'],
                    totals['rr_fees'],
                    totals['swf_fees'],
                    totals['twf_fees'],
                    totals['lab_fees'],
                    totals['sp_fees'],
                    totals['nss_fees'],
                    totals['dev_fees'],
                ])

                # reset
                min_fees_no = obj.fees_receipt_no
                max_fees_no = obj.fees_receipt_no
                totals = {k: 0 for k in FIELDS}
                current_date = obj.date

            else:
                max_fees_no = obj.fees_receipt_no

            for key in FIELDS:
                value = getattr(obj, key) or 0
                totals[key] += value
                grand_totals[key] += value   # ✅ accumulate

        # last row
        ws.append([
            current_date,
            f"{min_fees_no}-{max_fees_no}",
            totals['total_fees'],
            totals['tution_fees'],
            totals['admission_fees'],
            totals['id_fees'],
            totals['management_fees'],
            totals['lib_fees'],
            totals['assn_fees'],
            totals['rr_fees'],
            totals['swf_fees'],
            totals['twf_fees'],
            totals['lab_fees'],
            totals['sp_fees'],
            totals['nss_fees'],
            totals['dev_fees'],
        ])

    # ✅ EMPTY ROW
    ws.append([])

    # ✅ GRAND TOTAL ROW
    ws.append([
        'GRAND TOTAL',
        '',
        grand_totals['total_fees'],
        grand_totals['tution_fees'],
        grand_totals['admission_fees'],
        grand_totals['id_fees'],
        grand_totals['management_fees'],
        grand_totals['lib_fees'],
        grand_totals['assn_fees'],
        grand_totals['rr_fees'],
        grand_totals['swf_fees'],
        grand_totals['twf_fees'],
        grand_totals['lab_fees'],
        grand_totals['sp_fees'],
        grand_totals['nss_fees'],
        grand_totals['dev_fees'],
    ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime("%d-%m-%Y---%H-%M-%S")

    filename = f"day_history__{timestamp}.xlsx"

    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response


    wb.save(response)
    return response
@login_required
def student_history(request,roll_no,year):
    y=int(year)
    student_list=[]
    for i in History.objects.all():
        if i.student.roll_no2==roll_no and i.year==y:
            student_list.append(i)
    context={
        'student_list':student_list
        
    }
    return render(request,'student_history.html',context)
def print_receipt(request,receipt_no):
    hist=History.objects.get(pk=receipt_no)
    context={}
    context['hist']=hist
    return render(request,'print_reciept.html',context)   
def success(request):
    return render(request, 'success.html')


def abstract(request):
    m={} 
    total,collection,balance=0,0,0
    if request.method=="POST":
        academic_year=request.POST.get('academic_year')
        for i in Fees_Details.objects.filter(is_detained=False):
            if (i.academic_year.academic_year==academic_year or academic_year=="---") and (i.student.cancel_admission==False) and i.year!=0:
                if (i.student.dep,i.year) in m:
                    m[(i.student.dep,i.year)][0]+=i.total_fees 
                    m[(i.student.dep,i.year)][1]+=i.collection 
                    m[(i.student.dep,i.year)][2]+=i.balance 
                    m[(i.student.dep,i.year)][3]+=1
                else:
                    m[(i.student.dep,i.year)]=[i.total_fees,i.collection,i.balance,1] 
        l=[]
        for i in m:
            l.append([i[0],i[1],m[i][0],m[i][1],m[i][2],m[i][3]])
        l.sort()
        abst_list=[]
        c=1 
        total_students=0
        for i in l:
            obj=Abst(c,i[0],i[1],i[2],i[3],i[4],i[5])
            total+=i[2]
            collection+=i[3]
            balance+=i[4]
            total_students+=i[5]
            c+=1
            abst_list.append(obj) 
        abst_list.append(Abst("Total","","",total,collection,balance,total_students))
        year_list=Academic_Year.objects.all().order_by('-academic_year')
        context={'abst_list':abst_list,'year_list':year_list,'total':total,'collection':collection,'balance':balance,'academic_year':academic_year}
        return render(request,'abstract.html',context)
    else:
        m={}
        l=[]
        for i in Fees_Details.objects.filter(is_detained=False):
            if (i.student.cancel_admission==False) and i.year!=0:              
                if (i.student.dep,i.year) in m:
                        m[(i.student.dep,i.year)][0]+=i.total_fees 
                        m[(i.student.dep,i.year)][1]+=i.collection 
                        m[(i.student.dep,i.year)][2]+=i.balance
                        m[(i.student.dep,i.year)][3]+=1
                else:
                    m[(i.student.dep,i.year)]=[i.total_fees,i.collection,i.balance,1]
        for i in m:
            l.append([i[0],i[1],m[i][0],m[i][1],m[i][2],m[i][3]])
        l.sort()
        abst_list=[]
        c=1
        total_students=0
        for i in l: 
            obj=Abst(c,i[0],i[1],i[2],i[3],i[4],i[5])
            total+=i[2]
            collection+=i[3]
            balance+=i[4]
            total_students+=i[5]
            c+=1
            abst_list.append(obj) 
        abst_list.append(Abst("Total","","",total,collection,balance,total_students))
        year_list=Academic_Year.objects.all().order_by('-academic_year')
        context={'abst_list':abst_list,'year_list':year_list,'total':total,'collection':collection,'balance':balance}
        return render(request,'abstract.html',context)




def abstract2(request):
    m = {}
    total, collection, balance = 0, 0, 0

    # Initialize year-wise totals and counts
    year_totals = {1: [0, 0, 0], 2: [0, 0, 0], 3: [0, 0, 0]}  # [total_fees, collection, balance]
    lateral_totals = [0, 0, 0]  # [total_fees, collection, balance]

    # Count students for each category
   
    academic_year = None  # Initialize the academic_year variable
    academic_year = request.POST.get('academic_year')
    first_year_count=0 
    second_year_count=0 
    third_year_count=0 
    lateral_count=0
    if request.method == "POST":
        
        for entry in Fees_Details.objects.filter(student__Is_lateral=False,is_detained=False):
            if (entry.academic_year.academic_year == academic_year or academic_year == "---") and not entry.student.cancel_admission and entry.year!=0:
                if entry.year==1:
                    first_year_count+=1
                elif entry.year==2:
                    second_year_count+=1
                else:
                    third_year_count+=1

                key = (entry.student.dep, entry.year)

                # Update the main dictionary `m`
                if key in m:
                    m[key][0] += entry.total_fees
                    m[key][1] += entry.collection
                    m[key][2] += entry.balance
                    m[key][3] += 1
                else:
                    m[key] = [entry.total_fees, entry.collection, entry.balance, 1]

                year_totals[entry.year][0] += entry.total_fees
                year_totals[entry.year][1] += entry.collection
                year_totals[entry.year][2] += entry.balance
        for entry in Fees_Details.objects.filter(student__Is_lateral=True,is_detained=False):  # Correct field name
            if (entry.academic_year.academic_year == academic_year or academic_year == "---") and not entry.student.cancel_admission and entry.year!=0:
                if entry.year==2:
                    
                    lateral_totals[0] += entry.total_fees
                    lateral_totals[1] += entry.collection
                    lateral_totals[2] += entry.balance
                    lateral_count+=1
                else:
                    key = (entry.student.dep, entry.year)
                    print(entry.student.name,entry.student.roll_no2,entry.student.dep, entry.year,entry.student.roll_no)
                    if key in m:
                        m[key][0] += entry.total_fees
                        m[key][1] += entry.collection
                        m[key][2] += entry.balance
                        m[key][3] += 1
                    else:
                        m[key] = [entry.total_fees, entry.collection, entry.balance, 1]
                    third_year_count+=1
                    if entry.year in year_totals:
                        year_totals[entry.year][0] += entry.total_fees
                        year_totals[entry.year][1] += entry.collection
                        year_totals[entry.year][2] += entry.balance
                            
    else:
        first_year_count = Fees_Details.objects.filter(year=1, student__cancel_admission=False,is_detained=False).count()
        lateral_count = Fees_Details.objects.filter(student__Is_lateral=True, student__cancel_admission=False,is_detained=False).count()
        second_year_count = Fees_Details.objects.filter(year=2, student__Is_lateral=False, student__cancel_admission=False,is_detained=False).count()
        third_year_count = Fees_Details.objects.filter(year=3, student__cancel_admission=False, student__Is_lateral=False,is_detained=False).count()

        # Get lateral fees data
        

    # Prepare data for rendering
    abst_list = [
        Abst(c, k[0], k[1], *v) for c, (k, v) in enumerate(sorted(m.items()), 1)
    ]

    # Calculate Grand Totals
    grand_total_fees = year_totals[1][0] + lateral_totals[0] + year_totals[2][0] + year_totals[3][0]
    grand_collection = year_totals[1][1] + lateral_totals[1] + year_totals[2][1] + year_totals[3][1]
    grand_balance = year_totals[1][2] + lateral_totals[2] + year_totals[2][2] + year_totals[3][2]

    # If no POST request has been made, set academic_year to "---"
    if academic_year is None:
        academic_year = "---"

    context = {
        'abst_list': abst_list,
        'year_list': Academic_Year.objects.all().order_by('-academic_year'),
        'first_year_total': year_totals[1][0],
        'first_year_collection': year_totals[1][1],
        'first_year_balance': year_totals[1][2],
        'first_year_count': first_year_count,  # Added count for 1st year
        'lateral_total': lateral_totals[0],
        'lateral_collection': lateral_totals[1],
        'lateral_balance': lateral_totals[2],
        'lateral_count': lateral_count,  # Added count for lateral
        'second_year_total': year_totals[2][0],
        'second_year_collection': year_totals[2][1],
        'second_year_balance': year_totals[2][2],
        'second_year_count': second_year_count,  # Added count for 2nd year
        'third_year_total': year_totals[3][0],
        'third_year_collection': year_totals[3][1],
        'third_year_balance': year_totals[3][2],
        'third_year_count': third_year_count,  # Added count for 3rd year
        'grand_total_fees': grand_total_fees,
        'grand_collection': grand_collection,
        'grand_balance': grand_balance,
        'academic_year': academic_year,
    }

    return render(request, 'abstract2.html', context)


def empty_row():
        return History(
            fees_receipt_no='.',student=None,year=' ',tution_fees='  ',
                                admission_fees= '  ',id_fees= '  ',management_fees=' ',lib_fees=' ',assn_fees=' ',
                                rr_fees=' ',swf_fees=' ',twf_fees=' ',
                                lab_fees=' ',sp_fees=' ',nss_fees=' ',dev_fees=' '
                                ,date= '  ',total_fees=" ")

@login_required
def history(request):

    academic_year = request.GET.get('academic_year') or '---'

    qs = History.objects.select_related('student')

    if academic_year != '---':
        qs = qs.filter(academic_year=academic_year)

    qs = qs.order_by('date', 'fees_receipt_no')

    # =========================
    # ✅ GRAND TOTAL (FULL DATA)
    # =========================
    grand_totals = {
        'tution_fees': 0,
        'admission_fees': 0,
        'id_fees': 0,
        'management_fees': 0,
        'lib_fees': 0,
        'assn_fees': 0,
        'rr_fees': 0,
        'swf_fees': 0,
        'twf_fees': 0,
        'lab_fees': 0,
        'sp_fees': 0,
        'nss_fees': 0,
        'dev_fees': 0,
        'total_fees': 0,
    }

    for obj in qs:
        for k in grand_totals:
            grand_totals[k] += getattr(obj, k) or 0

    # =========================
    # PAGINATION
    # =========================
    paginator = Paginator(qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    day_list2 = []

    prev_date = None
    totals = None

    def reset_totals(obj):
        return {
            'tution_fees': obj.tution_fees or 0,
            'admission_fees': obj.admission_fees or 0,
            'id_fees': obj.id_fees or 0,
            'management_fees': obj.management_fees or 0,
            'lib_fees': obj.lib_fees or 0,
            'assn_fees': obj.assn_fees or 0,
            'rr_fees': obj.rr_fees or 0,
            'swf_fees': obj.swf_fees or 0,
            'twf_fees': obj.twf_fees or 0,
            'lab_fees': obj.lab_fees or 0,
            'sp_fees': obj.sp_fees or 0,
            'nss_fees': obj.nss_fees or 0,
            'dev_fees': obj.dev_fees or 0,
            'total_fees': obj.total_fees or 0,
        }

    for i in page_obj:

        if prev_date is None:
            totals = reset_totals(i)
            prev_date = i.date
            day_list2.append(i)
            continue

        if i.date == prev_date:
            for k in totals:
                totals[k] += getattr(i, k) or 0
        else:
            day_list2.append(History(
                fees_receipt_no='Total',
                student=None,
                date='',
                **totals
            ))

            day_list2.append(empty_row())

            totals = reset_totals(i)
            prev_date = i.date

        day_list2.append(i)

    if page_obj:
        day_list2.append(History(
            fees_receipt_no='Total',
            student=None,
            date='',
            **totals
        ))
        day_list2.append(empty_row())

    context = {
        'day_list': day_list2,
        'page_obj': page_obj,
        'year_list': Academic_Year.objects.all().order_by('-academic_year'),
        'academic_year': academic_year,
        'grand_totals': grand_totals   # ✅ ADD THIS
    }

    return render(request, 'history.html', context)


@login_required
def export_history_excel(request):

    academic_year = request.GET.get('academic_year') or '---'

    qs = History.objects.select_related('student')

    if academic_year != '---':
        qs = qs.filter(academic_year=academic_year)

    qs = qs.order_by('date', 'fees_receipt_no')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "History"

    headers = [
        "Date", "Receipt_No", "Name", "Reg_No", "Total",
        "Tution", "Admission", "ID", "Mgmt",
        "Lib", "Assn", "RR", "SWF", "TWF",
        "Lab", "SP", "NSS", "Dev"
    ]
    ws.append(headers)

    # =========================
    # TOTAL STRUCTURE
    # =========================
    FIELDS = [
        'tution_fees','admission_fees','id_fees','management_fees',
        'lib_fees','assn_fees','rr_fees','swf_fees','twf_fees',
        'lab_fees','sp_fees','nss_fees','dev_fees','total_fees'
    ]

    def reset_totals():
        return {k: 0 for k in FIELDS}

    prev_date = None
    totals = reset_totals()

    # =========================
    # GRAND TOTAL (FIXED)
    # =========================
    grand_totals = reset_totals()

    for i in qs:

        # =========================
        # FIXED: GRAND TOTAL ALWAYS UPDATED
        # =========================
        for k in grand_totals:
            grand_totals[k] += getattr(i, k) or 0

        # =========================
        # FIRST ROW
        # =========================
        if prev_date is None:
            for k in totals:
                totals[k] = getattr(i, k) or 0
            prev_date = i.date

        # =========================
        # SAME DATE
        # =========================
        elif i.date == prev_date:
            for k in totals:
                totals[k] += getattr(i, k) or 0

        # =========================
        # DATE CHANGE → PRINT TOTAL
        # =========================
        else:
            ws.append([
                "", "Total", "", "",
                totals['total_fees'],
                totals['tution_fees'],
                totals['admission_fees'],
                totals['id_fees'],
                totals['management_fees'],
                totals['lib_fees'],
                totals['assn_fees'],
                totals['rr_fees'],
                totals['swf_fees'],
                totals['twf_fees'],
                totals['lab_fees'],
                totals['sp_fees'],
                totals['nss_fees'],
                totals['dev_fees'],
            ])

            ws.append([""] * 18)

            totals = reset_totals()
            for k in totals:
                totals[k] = getattr(i, k) or 0

            prev_date = i.date

        # =========================
        # NORMAL ROW
        # =========================
        ws.append([
            str(i.date),
            i.fees_receipt_no,
            i.student.name if i.student else "",
            i.student.roll_no2 if i.student else "",
            i.total_fees,
            i.tution_fees,
            i.admission_fees,
            i.id_fees,
            i.management_fees,
            i.lib_fees,
            i.assn_fees,
            i.rr_fees,
            i.swf_fees,
            i.twf_fees,
            i.lab_fees,
            i.sp_fees,
            i.nss_fees,
            i.dev_fees,
        ])

    # =========================
    # LAST GROUP TOTAL
    # =========================
    if qs.exists():
        ws.append([
            "", "Total", "", "",
            totals['total_fees'],
            totals['tution_fees'],
            totals['admission_fees'],
            totals['id_fees'],
            totals['management_fees'],
            totals['lib_fees'],
            totals['assn_fees'],
            totals['rr_fees'],
            totals['swf_fees'],
            totals['twf_fees'],
            totals['lab_fees'],
            totals['sp_fees'],
            totals['nss_fees'],
            totals['dev_fees'],
        ])

    # =========================
    # GRAND TOTAL ROW
    # =========================
    ws.append([])
    ws.append([
        "GRAND TOTAL", "", "", "",
        grand_totals['total_fees'],
        grand_totals['tution_fees'],
        grand_totals['admission_fees'],
        grand_totals['id_fees'],
        grand_totals['management_fees'],
        grand_totals['lib_fees'],
        grand_totals['assn_fees'],
        grand_totals['rr_fees'],
        grand_totals['swf_fees'],
        grand_totals['twf_fees'],
        grand_totals['lab_fees'],
        grand_totals['sp_fees'],
        grand_totals['nss_fees'],
        grand_totals['dev_fees'],
    ])

    # =========================
    # FILE NAME (IST TIMESTAMP)
    # =========================
    timestamp = datetime.now().strftime("%d-%m-%Y---%H-%M-%S")
    filename = f"academic_year_history__{timestamp}.xlsx"

    # =========================
    # RESPONSE
    # =========================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response

# =========================
# ✅ UI VIEW
# =========================
@login_required
def from_date_to_date_history(request):

    from_date = None
    to_date = None

    history_qs = History.objects.select_related('student').order_by('date', 'fees_receipt_no')

    if request.method == 'POST':
        from_date_str = request.POST.get('from_date')
        to_date_str = request.POST.get('to_date')

        if from_date_str and to_date_str:
            from_date = datetime.strptime(from_date_str, "%Y-%m-%d").date()
            to_date = datetime.strptime(to_date_str, "%Y-%m-%d").date()

            history_qs = history_qs.filter(date__range=[from_date, to_date])

    # ✅ -------- GRAND TOTAL (FULL FILTERED DATA) --------
    grand_totals = {k: 0 for k in FIELDS}

    for obj in history_qs:
        for key in FIELDS:
            grand_totals[key] += getattr(obj, key) or 0

    # ✅ PAGINATION
    paginator = Paginator(history_qs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    page_list = list(page_obj)

    # ✅ DISPLAY LIST (with daily totals)
    final_list = []

    if page_list:
        current_date = page_list[0].date

        totals = {k: 0 for k in FIELDS}

        for obj in page_list:

            if obj.date != current_date:
                # 👉 total row
                final_list.append(History(
                    fees_receipt_no='Total',
                    student=None,
                    date='',
                    **totals
                ))

                # 👉 empty row
                final_list.append(empty_row())

                totals = {k: 0 for k in FIELDS}
                current_date = obj.date

            # accumulate
            for key in FIELDS:
                totals[key] += getattr(obj, key) or 0

            final_list.append(obj)

        # last total
        final_list.append(History(
            fees_receipt_no='Total',
            student=None,
            date='',
            **totals
        ))

    context = {
        'day_list': final_list,
        'page_obj': page_obj,
        'from_date': request.POST.get('from_date', ''),
        'to_date': request.POST.get('to_date', ''),
        'grand_totals': grand_totals   # ✅ ADD
    }

    return render(request, 'from_date_to_date_history.html', context)


# =========================
# ✅ EXCEL EXPORT
# =========================
@login_required
def from_date_to_date_history_excel(request):

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')

    qs = History.objects.select_related('student').order_by('date', 'fees_receipt_no')

    if from_date and to_date:
        qs = qs.filter(date__range=[from_date, to_date])

    qs = list(qs)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "History"

    headers = [
        'Date','Receipt','Name','Reg No','Total',
        'Tution','Admission','ID','Management','Library',
        'Assn','RR','SWF','TWF','Lab','SP','NSS','Dev'
    ]
    ws.append(headers)

    # ✅ GRAND TOTAL INIT
    grand_totals = {k: 0 for k in FIELDS}

    if qs:
        current_date = qs[0].date
        totals = {k: 0 for k in FIELDS}

        for obj in qs:

            # date break
            if obj.date != current_date:
                ws.append([
                    '', 'Total', '', '',
                    totals['total_fees'],
                    totals['tution_fees'],
                    totals['admission_fees'],
                    totals['id_fees'],
                    totals['management_fees'],
                    totals['lib_fees'],
                    totals['assn_fees'],
                    totals['rr_fees'],
                    totals['swf_fees'],
                    totals['twf_fees'],
                    totals['lab_fees'],
                    totals['sp_fees'],
                    totals['nss_fees'],
                    totals['dev_fees'],
                ])

                ws.append([])

                totals = {k: 0 for k in FIELDS}
                current_date = obj.date

            # accumulate
            for key in FIELDS:
                value = getattr(obj, key) or 0
                totals[key] += value
                grand_totals[key] += value   # ✅ GRAND TOTAL

            # normal row
            ws.append([
                obj.date,
                obj.fees_receipt_no,
                obj.student.name if obj.student else '',
                obj.student.roll_no2 if obj.student else '',
                obj.total_fees,
                obj.tution_fees,
                obj.admission_fees,
                obj.id_fees,
                obj.management_fees,
                obj.lib_fees,
                obj.assn_fees,
                obj.rr_fees,
                obj.swf_fees,
                obj.twf_fees,
                obj.lab_fees,
                obj.sp_fees,
                obj.nss_fees,
                obj.dev_fees,
            ])

        # last daily total
        ws.append([
            '', 'Total', '', '',
            totals['total_fees'],
            totals['tution_fees'],
            totals['admission_fees'],
            totals['id_fees'],
            totals['management_fees'],
            totals['lib_fees'],
            totals['assn_fees'],
            totals['rr_fees'],
            totals['swf_fees'],
            totals['twf_fees'],
            totals['lab_fees'],
            totals['sp_fees'],
            totals['nss_fees'],
            totals['dev_fees'],
        ])

    # ✅ EMPTY ROW
    ws.append([])

    # ✅ GRAND TOTAL ROW
    ws.append([
        'GRAND TOTAL', '', '', '',
        grand_totals['total_fees'],
        grand_totals['tution_fees'],
        grand_totals['admission_fees'],
        grand_totals['id_fees'],
        grand_totals['management_fees'],
        grand_totals['lib_fees'],
        grand_totals['assn_fees'],
        grand_totals['rr_fees'],
        grand_totals['swf_fees'],
        grand_totals['twf_fees'],
        grand_totals['lab_fees'],
        grand_totals['sp_fees'],
        grand_totals['nss_fees'],
        grand_totals['dev_fees'],
    ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime("%d-%m-%Y---%H-%M-%S")

    filename = f"from_date_to_date__{timestamp}.xlsx"

    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response




@login_required 
def update_usn(request):
    if request.method=='POST':
        temp_usn=request.POST.get('usn')
        perm_usn=request.POST.get('usn2')
        obj=Student.objects.get(roll_no=temp_usn)
        obj.roll_no2=perm_usn 
        obj.save() 
        return HttpResponseRedirect(reverse('success')) 
    else:
        student_list=Student.objects.all()
        context={'student_list':student_list}
        return render(request,'update_usn.html',context)
    pass

def appln_fee(request):
    if request.method=='POST':
        amount=int(request.POST.get('amount'))
        name=request.POST.get('name')
        academic_year=request.POST.get('academic_year')
        receipt_no=request.POST.get('receipt_no')
        try:
            hist = Application_Fees.objects.get(pk=receipt_no)
            return render(request,'fee_reciept_exist.html')
        except:
            appln_obj=Application_Fees()
            appln_obj.name=name 
            appln_obj.amount=amount 
            appln_obj.academic_year=Academic_Year.objects.get(pk=academic_year)
            appln_obj.fees_receipt_no=receipt_no
            appln_obj.save()
            return HttpResponseRedirect(reverse('success'))
    else:
        context={
            'year_list':Academic_Year.objects.all().order_by('-academic_year')
        }
        return render(request,'appln_fee.html',context)
    
def appln_fee_total(request):
    if request.method=='POST':
        academic_year=request.POST.get('academic_year')
        appln_list=[]
        total=0
        for i in Application_Fees.objects.all():
            if i.academic_year.academic_year==academic_year:
                total+=i.amount 
                appln_list.append(i)
        appln_list.append(Application_Fees(name="Total",amount=total))
        context={
            'year_list':Academic_Year.objects.all().order_by('-academic_year'),
            'appln_list':appln_list
        }
        return render(request,'appln_fee_total.html',context)
    else:
        context={
            'year_list':Academic_Year.objects.all().order_by('-academic_year'),
            'appln_list':[]
        }
        return render(request,'appln_fee_total.html',context)
    
def update_date(request):
    date=Date.objects.get(pk="1")
    if request.method=='POST':
        date.date=request.POST.get('date')
        date.save()
    context={
        'date':Date.objects.get(pk="1")
    }
    return render(request,'update_date.html',context)


def admission_order(request):
    if request.method=='POST':
        year=request.POST.get('year')
        year=int(year)
        student_list = Student.objects.exclude(name='.').filter(admission_year=year)
        context={
            'student_list':student_list,
            'total': len(student_list)
        }
        return render(request, 'admission_order.html', context)
    else:
        context={
            
        }
        return render(request, 'admission_order.html', context)


def print_order(request,roll_no):
    hist=Student.objects.get(pk=roll_no)
    curr_academic_year = list(Fees_Details.objects.filter(student = hist))
    print(curr_academic_year)
    curr_academic_year = curr_academic_year[0].academic_year
    context={}
    
    for i in hist.history_set.all():
        temp=i
        break
    context={
        'hist':hist,
        'fees_receipt':temp,
        'curr_academic_year':curr_academic_year
    }
    return render(request,'print_order.html',context)


def pending_fees(request):
    total, collection, balance = 0, 0, 0
    count = 0
    academic_year = " "
    year_s = " "
    dep_s = " "
    alphabetical = request.POST.get('alphabetical') == 'on'  # Check if checkbox is selected

    if request.method == "POST":
        dept = request.POST.get('dept')
        year = request.POST.get('year')
        academic_year = request.POST.get('academic_year')
        student_list = []
        y = -1
        if year != "---":
            y = int(year)
        
        for i in Fees_Details.objects.filter(is_detained=False):
            if (i.year == y or y == -1) and (i.student.dep == dept or dept == "---") and \
               (i.academic_year.academic_year == academic_year or academic_year == "---") and \
               i.balance > 0 and not i.student.cancel_admission:
                
                total += i.total_fees
                collection += i.collection
                balance += i.balance
                count += 1
                student_list.append(i)
        
        # Sort student_list alphabetically if checkbox is selected
        if alphabetical:
            student_list.sort(key=lambda x: x.student.name.lower())  # Sort ignoring case

        year_list = Academic_Year.objects.all().order_by('-academic_year')
        context = {
            'student_list': student_list,  # No need to reverse if alphabetical sort applied
            'year_list': year_list,
            'total': total,
            'collection': collection,
            'balance': balance,
            'dept': dept,
            'academic_year': academic_year,
            'year': year,
            'alphabetical': alphabetical  # Pass checkbox state to the template
        }
        return render(request, 'pending_fees.html', context)
    else:
        year_list = Academic_Year.objects.all().order_by('-academic_year')
        context = {
            'student_list': [],
            'year_list': year_list,
            'total': total,
            'collection': collection,
            'balance': balance,
            'dept': "",
            'academic_year': "",
            'year': "",
            'alphabetical': False  # Checkbox not selected by default
        }
        return render(request, 'pending_fees.html', context)

    

@login_required
def cancelled_admissions(request):

    total,collection,balance=0,0,0
    count=0
    academic_year=" "
    year_s=" "
    dep_s=" "
    if request.method=="POST":
        academic_year=request.POST.get('academic_year')
        student_list=[]
        for i in Fees_Details.objects.all():
            if (i.academic_year.academic_year==academic_year or academic_year=="---") and (i.student.cancel_admission==True) and i.year==1:
                
                total+=i.total_fees
                collection+=i.collection
                balance+=i.balance
                count+=1
                history=[]
                for j in History.objects.filter(student=i.student).filter(year=i.year):
                   history.append(Collection(j.fees_receipt_no,j.date,j.total_fees))
                student_list.append(Fees_Collection(i,history))

                
        year_list=Academic_Year.objects.all().order_by('-academic_year')
        context={'student_list':student_list[::-1],'total':total,'collection':collection,'balance':balance,'academic_year':academic_year, 'year_list': year_list}
        return render(request,'cancelled_admissions.html',context)
    else:
        context={
            'student_list':[],
            'total':"",'collection':"",'balance':"",'academic_year':"",
            'year_list': Academic_Year.objects.all().order_by('-academic_year')
        }
        return render(request,'cancelled_admissions.html',context)
        pass



@login_required
def student_details(request):
    academic_year_id = request.GET.get('academic_year')

    students_qs = Student.objects.all().select_related()

    academic_year = None

    if academic_year_id:
        try:
            academic_year = Academic_Year.objects.get(pk=academic_year_id)
            students_qs = students_qs.filter(
                fees_details__academic_year=academic_year
            ).distinct()
        except Academic_Year.DoesNotExist:
            students_qs = Student.objects.none()

    # 🔥 Pagination (DB-level slicing happens here)
    paginator = Paginator(students_qs, 50)  # 50 per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    academic_years = Academic_Year.objects.all().order_by('-academic_year')

    student_details = []
    for student in page_obj:
        fee_detail = Fees_Details.objects.filter(
            student=student,
            academic_year=academic_year
        ).first()

        student_details.append({
            'student': student,
            'year': fee_detail.year if fee_detail else None,
        })

    context = {
        'academic_years': academic_years,
        'students': student_details,
        'page_obj': page_obj,
        'academic_year': academic_year_id,
    }

    return render(request, 'student_details.html', context)



@login_required
def export_students_excel(request):
    academic_year_id = request.GET.get('academic_year')

    students = Student.objects.all()

    academic_year = None
    if academic_year_id:
        academic_year = Academic_Year.objects.filter(pk=academic_year_id).first()
        if academic_year:
            students = students.filter(
                fees_details__academic_year=academic_year
            ).distinct()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    # Header
    headers = [
        "Roll No", "Name", "Year", "Gender", "Department",
        "Admission Year", "Category", "Sub Category",
        "Phone", "Parent Name", "Parent Phone",
        "Application No", "Merit No", "Is Lateral"
    ]
    ws.append(headers)

    for student in students.iterator():  # 🔥 memory efficient
        fee = Fees_Details.objects.filter(
            student=student,
            academic_year=academic_year
        ).first()

        ws.append([
            student.roll_no2,
            student.name,
            fee.year if fee else "",
            student.get_gender_display(),
            student.dep,
            student.admission_year,
            student.category,
            student.sub_category,
            student.student_phone_number,
            student.parent_name,
            student.parent_phone_number,
            student.application_number,
            student.merit_no,
            student.Is_lateral,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=students.xlsx'

    wb.save(response)
    return response

    
    
def sub_category(request):
    m={} 
    student_count=0
    if request.method=="POST":
        academic_year=request.POST.get('academic_year')
        for i in Fees_Details.objects.all():
            if (i.academic_year.academic_year==academic_year):
                if i.student.sub_category in m:
                    m[i.student.sub_category]+=1
                else:
                    m[i.student.sub_category]=1
                student_count+=1
        l=[]
        for i in m:
            obj=Sub_Category(i,m[i])
            l.append(obj)
        l.append(Sub_Category("Total",student_count))
        year_list=Academic_Year.objects.all().order_by('-academic_year')
        context={'abst_list':l,'year_list':year_list,'academic_year':academic_year}
        return render(request,'sub_category.html',context)
    else:
        year_list=Academic_Year.objects.all().order_by('-academic_year')
        context={'abst_list':[],'year_list':year_list}
        return render(request,'sub_category.html',context)
    
class SC_ST_STATS:
    def __init__(self,name,year,usn,cat,gender):
        self.name=name 
        self.year=year 
        self.usn=usn
        self.cat=cat
        if gender=='M':
            self.gender="Boys"
        elif gender=='F':
            self.gender='Girls'
        else:
            self.gender=gender

def sc_st_stats(request):
    year_list=Academic_Year.objects.all().order_by('-academic_year')
    
    if request.method=='POST':
        sc_st_list=[]
        academic_year=request.POST.get('academic_year')
        m={}
        for i in Fees_Details.objects.filter(student__cancel_admission=False):
            if (i.academic_year.academic_year==academic_year and (i.student.category=="SC" or i.student.category=="ST")):
                if (i.student.category,i.student.gender,i.year) in m:
                    m[(i.student.category,i.student.gender,i.year)]+=1 
                else:
                    m[(i.student.category,i.student.gender,i.year)]=1 
                sc_st_list.append(SC_ST_STATS(i.student.name,i.year,i.student.roll_no2,i.student.category,i.student.gender))
        for i in m:
            sc_st_list.append(SC_ST_STATS("",i[2],"",i[0],i[1]))

        sc_st_list.sort(key=lambda x:(x.cat,x.gender,x.year,x.name))
        sc_st_list2=[]
        for i in sc_st_list:
            if i.name=="" and i.usn=="":
                i.name=i.cat 
                i.usn=i.gender 
                temp=""
                if i.gender=='Boys':
                    temp="M"
                else:
                    temp='F'
                i.year=m[(i.cat,temp,i.year)]
                sc_st_list2.append(i)
                sc_st_list2.append(SC_ST_STATS("Name","Year","Reg No","",""))
            else:
                sc_st_list2.append(i)

        context={'sc_st_list':sc_st_list2,'year_list':year_list,'academic_year':academic_year}
        return render(request,'sc_st_stats.html',context)
    else:
        sc_st_list=[]
        context={'sc_st_list':sc_st_list,'year_list':year_list}
        return render(request,'sc_st_stats.html',context)
    pass


def fees_structure_view(request):
    academic_year = None
    fees_structures = None

    if request.method == "POST":
        academic_year_value = request.POST.get("academic_year")

        if academic_year_value:
            academic_year = get_object_or_404(
                Academic_Year,
                academic_year=academic_year_value
            )

    if not academic_year:
        academic_year = Academic_Year.objects.order_by('-academic_year').first()

    if academic_year:
        fees_structures = Fees_Structure.objects.filter(
            academic_year=academic_year
        ).order_by('year', 'category', 'repeater', 'is_lateral')

    return render(request, 'fees_structure.html', {
        'fees_structures': fees_structures,
        'academic_year': academic_year,
        'all_years': Academic_Year.objects.all().order_by('-academic_year')
    })


def detention_view(request):
    student_list = []
    year_list = Academic_Year.objects.all().order_by('-academic_year')

    dept = request.POST.get('dept', '---')
    year = request.POST.get('year', '---')
    academic_year = request.POST.get('academic_year', '---')

    if request.method == "POST":

        y = int(year) if year != "---" else None

        queryset = Fees_Details.objects.select_related('student', 'academic_year')

        if y:
            queryset = queryset.filter(year=y)

        if dept != "---":
            queryset = queryset.filter(student__dep=dept)

        if academic_year != "---":
            queryset = queryset.filter(academic_year__academic_year=academic_year)

        # Only students not already detained
        queryset = queryset.filter(is_detained=False, student__cancel_admission=False)

        student_list = queryset

    return render(request, 'detention.html', {
        'student_list': student_list,
        'year_list': year_list,
        'dept': dept,
        'year': year,
        'academic_year': academic_year
    })


def add_detention(request):
    if request.method == "POST":
        roll_no = request.POST.get('roll_no')
        year = request.POST.get('year')
        academic_year = request.POST.get('academic_year')

        try:
            fees = Fees_Details.objects.select_related('student').get(
                student__roll_no2=roll_no,
                year=year,
                academic_year__academic_year=academic_year
            )

            # Create detention entry
            Detentions.objects.create(
                student=fees.student,
                year=fees.year,
                academic_year=fees.academic_year
            )

            # Update Fees_Details
            fees.is_detained = True
            fees.save()

            return JsonResponse({'status': 'success'})

        except Fees_Details.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Not found'})
        

# -----------------------------
# VIEW PAGE
# -----------------------------
def remove_detention_view(request):
    student_list = []
    year_list = Academic_Year.objects.all().order_by('-academic_year')

    dept = request.POST.get('dept', '')
    year = request.POST.get('year', '')
    academic_year = request.POST.get('academic_year', '')

    if request.method == "POST":

        queryset = Detentions.objects.select_related('student', 'academic_year')

        if year:
            queryset = queryset.filter(year=int(year))

        if dept:
            queryset = queryset.filter(student__dep=dept)

        if academic_year:
            queryset = queryset.filter(academic_year__academic_year=academic_year)

        student_list = queryset.order_by('student__roll_no2')

    return render(request, 'remove_detention.html', {
        'student_list': student_list,
        'year_list': year_list,
        'dept': dept,
        'year': year,
        'academic_year': academic_year
    })


# -----------------------------
# REMOVE DETENTION (mark only)
# -----------------------------
def mark_detention_removed(request):
    if request.method == "POST":

        roll_no = request.POST.get('roll_no')
        year = request.POST.get('year')
        academic_year = request.POST.get('academic_year')

        try:
            detention = Detentions.objects.select_related('student').get(
                student__roll_no2=roll_no,
                year=year,
                academic_year__academic_year=academic_year
            )

            if detention.is_detention_removed:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Already removed'
                })

            detention.is_detention_removed = True
            detention.save()

            return JsonResponse({'status': 'success'})

        except Detentions.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Detention not found'
            })

    return JsonResponse({'status': 'error'})


# -----------------------------
# RETAIN DETENTION (DELETE + RESET FEES)
# -----------------------------
def retain_detention_view(request):
    if request.method == "POST":

        roll_no = request.POST.get('roll_no')
        year = request.POST.get('year')
        academic_year = request.POST.get('academic_year')

        try:
            with transaction.atomic():

                detention = Detentions.objects.select_related('student').get(
                    student__roll_no2=roll_no,
                    year=year,
                    academic_year__academic_year=academic_year
                )

                student = detention.student

                # ✅ reset fees status
                Fees_Details.objects.filter(
                    student=student,
                    year=year,
                    academic_year__academic_year=academic_year
                ).update(is_detained=False)

                # ❌ delete detention record
                detention.delete()

            return JsonResponse({'status': 'success'})

        except Detentions.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Detention not found'
            })

    return JsonResponse({'status': 'error'})
        

@login_required
def delete_history(request):

    if request.method == "POST":

        receipt_no = request.POST.get('receipt_no')
        roll_no = request.POST.get('roll_no')
        year = request.POST.get('year')

        try:
            with transaction.atomic():

                # 1. Get history record
                history = History.objects.select_related('student').get(
                    fees_receipt_no=receipt_no
                )

                student = history.student

                # 2. Get matching Fees_Details
                fees = Fees_Details.objects.get(
                    student=student,
                    year=history.year,
                    academic_year__academic_year=history.academic_year
                )

                # 3. Reverse transaction
                fees.collection -= history.total_fees
                fees.balance += history.total_fees

                # safety check
                if fees.collection < 0:
                    fees.collection = 0

                fees.save()

                # 4. Delete history
                history.delete()

            return JsonResponse({'status': 'success'})

        except History.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'History not found'})

        except Fees_Details.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Fees record not found'})

    return JsonResponse({'status': 'error', 'message': 'Invalid request'})
